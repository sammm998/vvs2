"""Mangder, kanslighetstal och osakerhet - och leveransen till kalkylatorn.

Excel med tre blad enligt CLAUDE.md fas 8:

    Sammanstallning   en rad per beteckning eller system, med flaggor
    Per ror           en rad per strak, sa att varje meter gar att harleda
    Avvikelser        allt som inte gick in i mangden, med orsak

Matvarden och uppskattningar blandas ALDRIG utan markning (R3), och det som
inte kunde kopplas redovisas som egen rad i stallet for att gissas (R10).
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field


@dataclass
class QuantityRow:
    """En rad i mangdforteckningen."""

    label: str
    system: str | None
    dimension: str | None
    length_m: float
    verticals: int
    bends: int
    strands: int
    status: str = "unsplit"
    kind: str = "matvarde"        # matvarde | uppskattning
    label_source: str = "layer"   # anchor | propagated | layer | none
    flags: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "Beteckning": self.label,
            "System": self.system or "",
            "Dimension": self.dimension or "",
            "Langd_m": round(self.length_m, 1),
            "Vertikala_st": self.verticals,
            "Bojar_st": self.bends,
            "Strak_st": self.strands,
            "Status": self.status,
            "Typ": self.kind,
            "Beteckningskalla": self.label_source,
            "Flaggor": ", ".join(self.flags),
        }


def build_rows(result) -> list[QuantityRow]:
    """En rad per stilkluster, plus en rad for det som ligger i maskad zon.

    Sa lange beteckningarna inte ar lasta ar systemet den finaste niva som
    gar att redovisa. Raden markeras darfor med beteckningskalla 'layer' och
    dimensionen lamnas TOM - aldrig gissad.
    """
    measurement = result.triage.per_system_is_measurement
    rows: list[QuantityRow] = []
    for q in sorted(result.quantities, key=lambda q: -q.length_m):
        layer = q.layer or q.cluster_id
        rows.append(
            QuantityRow(
                label=layer.split("|")[-1],
                system=_system_from_layer(layer),
                dimension=None,
                length_m=q.length_m,
                verticals=q.verticals,
                bends=q.bends,
                strands=q.strands,
                kind="matvarde" if measurement else "uppskattning",
                label_source="layer",
                flags=list(q.flags),
            )
        )
    masked = result.masked_length_m
    if masked > 0 or result.masked_verticals:
        rows.append(
            QuantityRow(
                label="I MASKAD ZON - ingar ej i mangden",
                system=None,
                dimension=None,
                length_m=masked,
                verticals=result.masked_verticals,
                bends=0,
                strands=0,
                kind="matvarde" if measurement else "uppskattning",
                label_source="none",
                flags=["masked_zone", "reported_not_counted"],
            )
        )
    return rows


def _system_from_layer(layer: str) -> str | None:
    """Systembeteckningen som CAD-lagret sjalvt bar, om den gar att lasa ut.

    Sista faltet i lagernamnet ar konventionellt systemets kortnamn
    (``V-53BB-FE--S3-`` -> ``S3``). Gar det inte att lasa ut lamnas faltet
    tomt i stallet for att fyllas med en gissning.
    """
    tail = layer.split("|")[-1]
    fields = [f for f in tail.split("-") if f]
    return fields[-1] if fields else None


def sensitivity(result) -> dict:
    """Kanslighetstal: vad ar besluten varda?"""
    from .status import wall_zone_overlap

    plateaus = [cr.plateau_width for cr in result.chains.values()] or [0.0]
    bridged = sum(cr.bridged_length for cr in result.chains.values())
    total_pt = sum(s.length for s in result.net.strands) or 1.0
    return {
        "plataabredd_min": round(min(plateaus), 3),
        "andel_overbryggad_langd": round(bridged / total_pt, 4),
        "ror_i_vaggzon": round(wall_zone_overlap(result, result.zones), 4),
        "maskad_langd_m": round(result.masked_length_m, 1),
        "maskade_vertikala": result.masked_verticals,
        "skala_verifierad": result.scale.verified,
        "spar": result.triage.track,
        "urvalsmetod": result.selection.method,
    }


def export_excel(result, path: str) -> str:
    """Skriv leveransen. Tre blad, och flaggorna foljer med varje rad."""
    import openpyxl
    from openpyxl.styles import Alignment, Font

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()

    rows = build_rows(result)
    ws = wb.active
    ws.title = "Sammanstallning"
    _write_table(ws, [r.as_dict() for r in rows])
    _summary_block(ws, result, len(rows) + 3)

    ws2 = wb.create_sheet("Per ror")
    per_strand = []
    for s in sorted(result.net.strands, key=lambda s: -s.length):
        layer = result.styles.get(s.cluster_id).key.layer or s.cluster_id
        per_strand.append(
            {
                "Strak_id": s.id,
                "Beteckning": layer.split("|")[-1],
                "System": _system_from_layer(layer) or "",
                "Langd_m": round(result.scale.to_m(s.length), 2),
                "Punkter": len(s.points),
                "Nod_A": s.node_a,
                "Nod_B": s.node_b,
            }
        )
    _write_table(ws2, per_strand)

    ws3 = wb.create_sheet("Avvikelser")
    reasons: dict[str, int] = {}
    for b in result.selection.blocked:
        reasons[b.reason] = reasons.get(b.reason, 0) + 1
    dev = [
        {"Orsak": k, "Antal_banor": v, "Steg": "urval"}
        for k, v in sorted(reasons.items(), key=lambda kv: -kv[1])
    ]
    for f in result.flags + result.selection.flags + result.scale.flags:
        dev.append({"Orsak": f, "Antal_banor": "", "Steg": "flagga"})
    _write_table(ws3, dev)

    for sheet in wb:
        for col in sheet.columns:
            width = max((len(str(c.value or "")) for c in col), default=8)
            sheet.column_dimensions[col[0].column_letter].width = min(46, max(11, width + 2))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
    wb.save(path)
    return path


def _write_table(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["(inga rader)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])


def _summary_block(ws, result, start_row: int) -> None:
    from openpyxl.styles import Font

    s = sensitivity(result)
    ws.cell(row=start_row, column=1, value="KANSLIGHETSTAL OCH FORUTSATTNINGAR").font = Font(bold=True)
    items = [
        ("Ritning", result.drawing),
        ("Spar (R3)", s["spar"]),
        ("Skala", f"1:{result.scale.value:.0f}"),
        ("Skala verifierad geometriskt (R4)", "JA" if s["skala_verifierad"] else "NEJ - preliminar"),
        ("Skalkallor", ", ".join(x.name for x in result.scale.sources)),
        ("Urvalsmetod for rorstil", s["urvalsmetod"]),
        ("Tackning (R6)", f"{result.coverage:.4f}"),
        ("Total langd (m)", round(result.total_length_m, 1)),
        ("Langd i maskad zon, ej mangdad (m)", s["maskad_langd_m"]),
        ("Vertikala ror i maskad zon, ej raknade", s["maskade_vertikala"]),
        ("Andel langd som ar overbryggade glapp", s["andel_overbryggad_langd"]),
        ("Andel ror i vaggzon", s["ror_i_vaggzon"]),
        ("Minsta plataabredd vid sammanfogning", s["plataabredd_min"]),
    ]
    for i, (k, v) in enumerate(items, start=1):
        ws.cell(row=start_row + i, column=1, value=k)
        ws.cell(row=start_row + i, column=2, value=v)
