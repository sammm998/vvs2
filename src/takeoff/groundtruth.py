"""Facit-import.

Facit levereras som en Bluebeam-markeringsexport i Excel: en rad per
markering, med beteckning i ``Subject``, langd i ``Langd`` och antal vertikala
ror i ``Antal_VS``. Kolumnnamnen varierar mellan exporter, sa de loses upp mot
aliaslistor och den *upplosta mappningen skrivs ut* - den hardkodas aldrig
tyst.

Nar den uppmarkta PDF:en finns med kan aven markeringarnas geometri lasas.
Den anvands enbart for utvardering, aldrig av matmotorn.

Omimport raderar tidigare facit for samma ritning i stallet for att lagga
till.
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field

ALIASES = {
    "label": ("beteckning", "label", "benamning", "benämning", "subject", "ämne", "amne"),
    "length": ("langd", "längd", "length", "m", "matt", "mått"),
    "count": ("antal", "antal_vs", "vertikala", "st", "count"),
    "layer": ("lager", "layer", "skikt"),
    "document": ("document", "dokument", "ritning", "drawing"),
    "unit": ("unit", "enhet"),
}

VERTICAL_SUFFIX = re.compile(r"\s+vertikal(t)?$", re.IGNORECASE)
LABEL_GRAMMAR = re.compile(r"^([A-ZÅÄÖ]+\d*)-([A-Z0-9]+)(?:-(\d+))?$", re.IGNORECASE)


@dataclass
class GroundTruthRow:
    drawing: str
    label: str
    system: str | None
    material: str | None
    dimension: str | None
    length_m: float | None
    vertical_count: float | None
    layer: str | None
    is_vertical: bool


@dataclass
class GroundTruth:
    drawing: str
    rows: list[GroundTruthRow]
    column_mapping: dict[str, str]
    unmapped_columns: list[str] = field(default_factory=list)
    geometry: list[dict] = field(default_factory=list)

    def totals(self) -> dict:
        by_label: dict[str, float] = {}
        verticals: dict[str, float] = {}
        for r in self.rows:
            if r.length_m:
                by_label[r.label] = by_label.get(r.label, 0.0) + r.length_m
            if r.vertical_count:
                verticals[r.label] = verticals.get(r.label, 0.0) + r.vertical_count
        return {
            "total_length_m": round(sum(by_label.values()), 2),
            "total_verticals": round(sum(verticals.values()), 2),
            "labels": {k: round(v, 2) for k, v in sorted(by_label.items(), key=lambda x: -x[1])},
            "verticals": {k: v for k, v in sorted(verticals.items(), key=lambda x: -x[1])},
        }


def _norm(h) -> str:
    return re.sub(r"[^a-z0-9åäö_]", "", str(h or "").strip().lower())


def _resolve_columns(headers: list) -> tuple[dict[str, int], list[str]]:
    norm = [_norm(h) for h in headers]
    mapping: dict[str, int] = {}
    for field_name, aliases in ALIASES.items():
        for i, h in enumerate(norm):
            if h in aliases and field_name not in mapping:
                mapping[field_name] = i
                break
        if field_name not in mapping:
            for i, h in enumerate(norm):
                if h and any(h.startswith(a) or a in h for a in aliases) and i not in mapping.values():
                    mapping[field_name] = i
                    break
    used = set(mapping.values())
    unmapped = [str(headers[i]) for i in range(len(headers)) if i not in used and headers[i]]
    return mapping, unmapped


def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_label(label: str) -> tuple[str | None, str | None, str | None]:
    """SYSTEM-TYP[-DIMENSION]. Rader utan dimension ar GILTIGA."""
    m = LABEL_GRAMMAR.match(label.strip())
    if not m:
        return (None, None, None)
    return (m.group(1).upper(), m.group(2).upper(), m.group(3))


def load_excel(path: str, drawing: str | None = None, verbose: bool = True) -> GroundTruth:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    mapping, unmapped = _resolve_columns(headers)
    if verbose:
        print("Facit-kolumner i filen:")
        for h in headers:
            print(f"    {h!r}")
        print("Upplost mappning (bekrafta innan den anvands i leverans):")
        for k, i in sorted(mapping.items()):
            print(f"    {k:9s} <- {headers[i]!r}")
        if unmapped:
            print(f"    oanvanda kolumner: {unmapped}")

    if "label" not in mapping:
        raise ValueError("hittade ingen beteckningskolumn i facitfilen")

    rows: list[GroundTruthRow] = []
    doc_col = mapping.get("document")
    for raw in ws.iter_rows(min_row=2, values_only=True):
        label_raw = raw[mapping["label"]]
        if not label_raw:
            continue
        label_raw = str(label_raw).strip()
        is_vert = bool(VERTICAL_SUFFIX.search(label_raw))
        label = VERTICAL_SUFFIX.sub("", label_raw).strip()
        system, material, dim = parse_label(label)
        drw = drawing or (str(raw[doc_col]) if doc_col is not None and raw[doc_col] else "")
        rows.append(
            GroundTruthRow(
                drawing=drw,
                label=label,
                system=system,
                material=material,
                dimension=dim,
                length_m=_num(raw[mapping["length"]]) if "length" in mapping else None,
                vertical_count=_num(raw[mapping["count"]]) if "count" in mapping else None,
                layer=str(raw[mapping["layer"]]) if "layer" in mapping and raw[mapping["layer"]] else None,
                is_vertical=is_vert,
            )
        )
    drw = drawing or (rows[0].drawing if rows else "")
    return GroundTruth(drawing=drw, rows=rows, column_mapping={k: str(headers[i]) for k, i in mapping.items()}, unmapped_columns=unmapped)


def load_markup_geometry(pdf_path: str, page_number: int = 0) -> list[dict]:
    """Las facitets markeringsgeometri ur en uppmarkt PDF.

    Endast for utvardering. Matmotorn ser aldrig denna fil.
    """
    import pymupdf

    from .normalize import page_transform

    doc = pymupdf.open(pdf_path)
    page = doc[page_number]
    # R8: aven facitgeometrin maste genom sidans transform. Annoteringarnas
    # vertices ligger i raa PDF-koordinater precis som banorna, och pa en
    # ritning med /Rotate hamnar de i en annan rymd an matningen om steget
    # hoppas over. Pa en oroterad ritning ar det identitet och syns aldrig.
    tf = page_transform(page)
    out: list[dict] = []
    for a in page.annots() or ():
        kind = a.type[1]
        if kind not in ("PolyLine", "Polygon"):
            continue
        verts = [tf.apply(tuple(v)) for v in (a.vertices or [])]
        if not verts:
            continue
        label_raw = (a.info.get("subject") or "").strip()
        is_vert = bool(VERTICAL_SUFFIX.search(label_raw)) or kind == "Polygon"
        out.append(
            {
                "kind": kind,
                "label": VERTICAL_SUFFIX.sub("", label_raw).strip(),
                "vertices": verts,
                "is_vertical": is_vert,
                "length_pt": sum(math.dist(verts[i], verts[i + 1]) for i in range(len(verts) - 1)),
                "rect": tuple(tf.apply_rect(a.rect)),
            }
        )
    doc.close()
    return out
