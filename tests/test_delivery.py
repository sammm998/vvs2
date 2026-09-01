"""Leveransen: mangdforteckningen far aldrig blanda matvarde och gissning."""

import os

from tests.conftest import needs_drawing


@needs_drawing
def test_dimension_is_never_invented(result):
    """Utan last beteckning ar dimensionen OKAND, inte gissad (R10).

    Lagret ger systemet men inte dimensionen. Att fylla i den fran narmaste
    rad vore precis den gissning CLAUDE.md forbjuder.
    """
    from takeoff import quantify

    for row in quantify.build_rows(result):
        if row.label_source == "layer":
            assert row.dimension is None


@needs_drawing
def test_measurement_and_estimate_are_labelled(result):
    """R3: ett spar B-resultat far aldrig se ut som ett matvarde."""
    from takeoff import quantify

    expected = "matvarde" if result.triage.track == "A" else "uppskattning"
    for row in quantify.build_rows(result):
        assert row.kind == expected


@needs_drawing
def test_masked_zone_has_its_own_row(result):
    """Det som maskas bort ska SYNAS i leveransen, inte forsvinna."""
    from takeoff import quantify

    rows = quantify.build_rows(result)
    if result.masked_length_m > 0:
        masked = [r for r in rows if "masked_zone" in r.flags]
        assert len(masked) == 1
        assert masked[0].length_m == result.masked_length_m
        assert "reported_not_counted" in masked[0].flags


@needs_drawing
def test_rows_sum_to_the_reported_total(result):
    """Summan av raderna maste vara totalen. Inget far tappas pa vagen."""
    from takeoff import quantify

    rows = quantify.build_rows(result)
    counted = sum(r.length_m for r in rows if "masked_zone" not in r.flags)
    assert abs(counted - result.total_length_m) < 1e-6


@needs_drawing
def test_excel_delivery_has_three_sheets(result, tmp_path):
    import openpyxl

    from takeoff import quantify

    path = quantify.export_excel(result, str(tmp_path / "m.xlsx"))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Sammanstallning", "Per ror", "Avvikelser"]
    assert wb["Per ror"].max_row > 1


@needs_drawing
def test_sensitivity_numbers_are_present(result):
    """Kanslighetstalen ska alltid folja med - de sager vad besluten ar varda."""
    from takeoff import quantify

    s = quantify.sensitivity(result)
    for key in ("plataabredd_min", "andel_overbryggad_langd", "ror_i_vaggzon",
                "maskad_langd_m", "skala_verifierad", "spar", "urvalsmetod"):
        assert key in s
